"""
Fungsi: FastAPI router untuk validasi format dokumen DOCX proposal mahasiswa.

Endpoint:
  POST /run                   — validasi satu dokumen (synchronous, hasil langsung dikembalikan)
  POST /bulk                  — validasi banyak dokumen (async, kembalikan session_id, proses di background)
  GET  /sessions/{id}         — cek status + hasil validasi bulk
  POST /summarize             — ringkasan naratif LLM dari daftar issue (satu dokumen)
  GET  /export/{session_id}   — generate Excel ringkasan LLM semua dokumen dalam bulk session

Digunakan oleh: backend/src/routes/pkm.routes.js (sebagai proxy)
Keyword: automated document validation
"""
import asyncio
import io
import json
import logging
import os
import tempfile
from datetime import datetime, timezone
from pathlib import Path

from fastapi import APIRouter, BackgroundTasks, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from services.database import get_supabase
from services.job_processor import build_validation_dict, process_bulk_session, save_temp_file

logger = logging.getLogger(__name__)

MAX_DOCX_SIZE  = 20 * 1024 * 1024  # 20 MB
MAX_BULK_ITEMS = 20


class SummarizeRequest(BaseModel):
    issues: list[dict] = []
    schema_name: str | None = None

router = APIRouter()


def _utcnow() -> str:
    return datetime.now(timezone.utc).isoformat()


# ─── Helper bersama: ambil metadata validasi dari Supabase ───────────────────

async def _fetch_metadata(schema_id: str, tahun: str) -> dict:

    supabase = get_supabase()

    proj = (
        supabase.table("projects")
        .select("id")
        .eq("skema", schema_id)
        .eq("tahun", tahun)
        .order("created_at", desc=True)
        .limit(1)
        .execute()
    )
    if not proj.data:
        raise HTTPException(
            status_code=404,
            detail=f"Data referensi untuk {schema_id} tahun {tahun} belum tersedia.",
        )

    meta_row = (
        supabase.table("document_metadata")
        .select("payload")
        .eq("project_id", proj.data[0]["id"])
        .limit(1)
        .execute()
    )
    if not meta_row.data:
        raise HTTPException(
            status_code=404,
            detail="Metadata format dokumen belum tersedia. Pastikan pipeline ekstraksi sudah selesai.",
        )

    payload = meta_row.data[0]["payload"]
    if isinstance(payload, str):
        payload = json.loads(payload)
    return payload


# ─── POST /run — validasi satu dokumen ───────────────────────────────────────

@router.post("/run")
async def run_validation(
    schema_id: str = Form(...),
    tahun: str = Form(...),
    file: UploadFile = File(...),
):

    schema_id = schema_id.strip().upper()
    tahun     = tahun.strip()

    if not (file.filename or "").lower().endswith(".docx"):
        raise HTTPException(status_code=400, detail="File harus berformat DOCX (.docx).")

    if not tahun:
        raise HTTPException(status_code=400, detail="Parameter tahun tidak boleh kosong.")

    try:
        payload = await _fetch_metadata(schema_id, tahun)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Gagal mengambil data dari database: {str(e)}")

    content = await file.read()
    if len(content) > MAX_DOCX_SIZE:
        raise HTTPException(status_code=400, detail="Ukuran file melebihi batas maksimum 20 MB.")

    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name

        from model_ai.validation import validate_document  # noqa: PLC0415
        result = validate_document(tmp_path, metadata_dict=payload)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Gagal menjalankan validasi dokumen: {str(e)}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)

    result_dict = build_validation_dict(result)


    try:
        supabase = get_supabase()
        now = _utcnow()

        session_row = (
            supabase.table("validation_sessions")
            .insert({
                "type":            "single",
                "status":          "completed",
                "total_items":     1,
                "completed_items": 1,
                "created_at":      now,
                "updated_at":      now,
            })
            .execute()
        )

        if session_row.data:
            session_id = session_row.data[0]["id"]
            supabase.table("validation_results").insert({
                "session_id":  session_id,
                "position":    0,
                "file_name":   file.filename or "dokumen.docx",
                "schema_id":   schema_id,
                "tahun":       tahun.strip(),
                "status":      "completed",
                "result":      result_dict,
                "created_at":  now,
                "updated_at":  now,
            }).execute()
    except Exception as db_exc:

        logger.warning("Gagal menyimpan hasil validasi ke DB: %s", db_exc)

    return result_dict


# ─── POST /summarize — ringkasan naratif LLM dari issue ──────────────────────

@router.post("/summarize")
async def summarize_validation(body: SummarizeRequest):

    if not body.issues:
        return {"summary": "", "generated_at": _utcnow()}

    try:
        from model_ai.validation.summarizer import summarize_issues  # noqa: PLC0415
        summary = summarize_issues(body.issues, body.schema_name)
    except Exception as e:
        raise HTTPException(
            status_code=503,
            detail=f"Gagal membuat ringkasan: {str(e)}",
        )

    return {"summary": summary, "generated_at": _utcnow()}


# ─── GET /export/{session_id} — Excel ringkasan LLM per dokumen ─────────────

def _extract_ketua(file_name: str) -> str:

    return Path(file_name).stem.split("_")[0]


@router.get("/export/{session_id}")
async def export_bulk_summary(session_id: str, schema_name: str | None = None):

    import openpyxl  # noqa: PLC0415

    supabase = get_supabase()

    results_row = (
        supabase.table("validation_results")
        .select("position, file_name, status, result, error_message")
        .eq("session_id", session_id)
        .order("position", desc=False)
        .execute()
    )
    if not results_row.data:
        raise HTTPException(
            status_code=404,
            detail="Session tidak ditemukan atau belum ada hasil validasi.",
        )

    from model_ai.validation.summarizer import summarize_issues  # noqa: PLC0415

    async def _summarize_item(item: dict) -> tuple[str, str]:
        ketua = _extract_ketua(item.get("file_name") or "unknown")

        result_raw = item.get("result") or {}
        if isinstance(result_raw, str):
            try:
                result_raw = json.loads(result_raw)
            except Exception:
                result_raw = {}
        issues = result_raw.get("issues") or [] if isinstance(result_raw, dict) else []

        if issues and item.get("status") == "completed":
            try:
                summary = await asyncio.to_thread(summarize_issues, issues, schema_name)
            except Exception as exc:
                summary = f"[Gagal generate ringkasan: {exc}]"
        else:
            summary = ""

        return ketua, summary

    rows_with_summaries: list[tuple[str, str]] = await asyncio.gather(
        *[_summarize_item(item) for item in results_row.data]
    )

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ringkasan Validasi"

    ws.append(["Nama Pemilik Proposal", "Ringkasan Kekurangan"])

    from openpyxl.styles import Font  # noqa: PLC0415
    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 80

    for ketua, summary in rows_with_summaries:
        ws.append([ketua, summary])

    from openpyxl.styles import Alignment  # noqa: PLC0415
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ── Serialize ke BytesIO dan kembalikan sebagai StreamingResponse ─────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    safe_session = session_id[:8] if len(session_id) >= 8 else session_id
    filename = f"ringkasan-validasi-{safe_session}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── POST /bulk — validasi banyak dokumen (async) ────────────────────────────

@router.post("/bulk")
async def run_bulk_validation(request: Request, background_tasks: BackgroundTasks):

    try:
        form = await request.form()
        count = int(form.get("count") or 0)
    except Exception:
        raise HTTPException(status_code=400, detail="Format form data tidak valid.")

    if count < 1:
        raise HTTPException(status_code=400, detail="Minimal satu dokumen harus diupload.")

    if count > MAX_BULK_ITEMS:
        raise HTTPException(
            status_code=400,
            detail=f"Maksimal {MAX_BULK_ITEMS} dokumen per batch.",
        )


    items_bytes: list[dict] = []
    for i in range(count):
        schema_id = str(form.get(f"schema_id_{i}") or "").strip().upper()
        tahun     = str(form.get(f"tahun_{i}")     or "").strip()
        upload    = form.get(f"file_{i}")

        if not schema_id:
            raise HTTPException(status_code=400, detail=f"schema_id dokumen ke-{i + 1} kosong.")
        if not tahun:
            raise HTTPException(status_code=400, detail=f"tahun dokumen ke-{i + 1} kosong.")
        if upload is None:
            raise HTTPException(status_code=400, detail=f"File dokumen ke-{i + 1} tidak ditemukan.")

        filename = getattr(upload, "filename", "") or f"dokumen_{i + 1}.docx"
        if not filename.lower().endswith(".docx"):
            raise HTTPException(
                status_code=400,
                detail=f"Dokumen ke-{i + 1} ({filename}) harus berformat DOCX.",
            )

        content = await upload.read()
        if len(content) > MAX_DOCX_SIZE:
            raise HTTPException(
                status_code=400,
                detail=f"Dokumen ke-{i + 1} ({filename}) melebihi batas maksimum 20 MB.",
            )

        items_bytes.append({
            "position":  i,
            "file_name": filename,
            "schema_id": schema_id,
            "tahun":     tahun,
            "content":   content,
        })


    supabase = get_supabase()
    now = _utcnow()

    session_row = (
        supabase.table("validation_sessions")
        .insert({
            "type":            "bulk",
            "status":          "pending",
            "total_items":     count,
            "completed_items": 0,
            "created_at":      now,
            "updated_at":      now,
        })
        .execute()
    )
    if not session_row.data:
        raise HTTPException(status_code=500, detail="Gagal membuat validation session di database.")

    session_id = session_row.data[0]["id"]


    result_rows = [
        {
            "session_id": session_id,
            "position":   item["position"],
            "file_name":  item["file_name"],
            "schema_id":  item["schema_id"],
            "tahun":      item["tahun"],
            "status":     "pending",
            "created_at": now,
            "updated_at": now,
        }
        for item in items_bytes
    ]
    supabase.table("validation_results").insert(result_rows).execute()


    for item in items_bytes:
        save_temp_file(session_id, item["position"], item["content"])


    items_meta = [
        {
            "position":  item["position"],
            "file_name": item["file_name"],
            "schema_id": item["schema_id"],
            "tahun":     item["tahun"],
        }
        for item in items_bytes
    ]
    background_tasks.add_task(process_bulk_session, session_id, items_meta)

    return {"session_id": session_id}


# ─── GET /sessions/{session_id} — status + hasil validasi bulk ───────────────

@router.get("/sessions/{session_id}")
async def get_session_status(session_id: str):

    supabase = get_supabase()

    session_row = (
        supabase.table("validation_sessions")
        .select("id, type, status, total_items, completed_items, created_at, updated_at")
        .eq("id", session_id)
        .limit(1)
        .execute()
    )
    if not session_row.data:
        raise HTTPException(status_code=404, detail="Session tidak ditemukan.")

    session = session_row.data[0]

    results_row = (
        supabase.table("validation_results")
        .select("id, position, file_name, schema_id, tahun, status, result, error_message")
        .eq("session_id", session_id)
        .order("position", desc=False)
        .execute()
    )

    return {
        "id":              session["id"],
        "type":            session["type"],
        "status":          session["status"],
        "total_items":     session["total_items"],
        "completed_items": session["completed_items"],
        "created_at":      session["created_at"],
        "updated_at":      session["updated_at"],
        "items": [
            {
                "id":            item["id"],
                "position":      item["position"],
                "file_name":     item["file_name"],
                "schema_id":     item["schema_id"],
                "tahun":         item["tahun"],
                "status":        item["status"],
                "result":        item["result"],
                "error_message": item["error_message"],
            }
            for item in results_row.data
        ],
    }
